#!/usr/bin/env python3

import sys
import shutil
from openpyxl import load_workbook
import datetime
import calendar
import lxml.etree as ET
import tdreader
import re
import os



def usage():
    print("Usage:   td2cambio <year>-<month> <xml-file>")
    print("Example: td2cambio 2020-02 Export.xml")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        usage()
        exit(1)

    # Year and month from argv
    year, month = sys.argv[1].split("-")
    year = int(year)
    month = int(month)
    month_name = calendar.month_name[month]

    # TBD check that year and month are integers

    # Create new file with correct name
    file_name = "Konsulttidrapport-Responsive-"+sys.argv[1]+".xlsx"
    dir_of_script = os.path.dirname(os.path.realpath(__file__))
    shutil.copyfile(os.path.join(dir_of_script, "Cambio-mall.xlsx"), file_name)

    # Open workbook
    workbook = load_workbook(filename = file_name)
    sheet = workbook.active

    # The name of the activity in Timeduty should be the team followed by the
    # project number in parenthesis so that we can extract them automatically
    activity_name = "Java SDK Teamet (1234-1)"
    team = activity_name.split("(")[0].strip()
    projno = activity_name.split("(")[1].strip()[:-1]

    # Write month info and team
    sheet.cell(column=2, row=4, value=month_name+" "+str(year))
    sheet.cell(column=2, row=5, value=team)

    # Where to write dates and times?
    first_row = 8
    time_column = 2

    # Update dates for the rows
    date = datetime.datetime(year, month, 1)
    for row in range(first_row, first_row+31):
        sheet.cell(column=1, row=row, value=date)
        date = date + datetime.timedelta(days=1)

    # Fill project number
    sheet.cell(column=2, row=7, value=projno)

    # and hours from the xml-file
    input_file = sys.argv[2]
    indata = ET.parse(input_file)

    from_date, to_date, time_rows, expense_rows = tdreader.extract_data_from_xml(indata)

    # TODO check that from_date & to_date matches the month given as argument


    # Get all cambio project activity time registrations
    registrations = tdreader.filter_registrations_for_client("Cambio", time_rows)
    if len(registrations) == 0:
        print("ERROR: No registrations found in '{}'".format(input_file))
        sys.exit(1)

    # TODO we should probably check that there is only one user in the registrations
    # or possibly generate one report per person
    name = tdreader.get_name(registrations[0])
    sheet.cell(column=2, row=2, value=name)     # Write name in name cell

    # TODO we should probably ensure that there is only one project number
    # or place hours for different project numbers in different columns
    activity_name = tdreader.get_activity(registrations[0])
    activity_number = re.sub(r'.*\((.*)\).*', r'\1', activity_name)
    sheet.cell(column=2, row=7, value=activity_number)

    # Get hours from the XML-file and input in corresponding cells
    for registration in registrations:
        y, m, d = tdreader.get_date(registration).split('-')
        time = float(tdreader.get_time(registration))
        sheet.cell(column=time_column, row=first_row+int(d)-1, value=time)

    # Close and exit
    workbook.save(file_name)
